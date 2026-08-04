"""
Formatted multi-sheet Excel export utility using openpyxl.
"""

import io
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.models.email import CanonicalEmailResult
from src.models.scan import ScanError, ScanStats

# Style objects are immutable in practice here and openpyxl copies them onto each
# cell, so building them once beats reconstructing per row of a large export.
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
LINK_FONT = Font(name="Calibri", size=11, color="0563C1", underline="single")


def generate_excel_bytes(
    results: List[CanonicalEmailResult],
    stats: ScanStats,
    errors: List[ScanError]
) -> bytes:
    """
    Generates a formatted XLSX workbook bytes buffer with 3 sheets:
    'Emails', 'Scan Summary', and 'Errors'.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Header styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    # 1. Emails Sheet
    ws_emails = wb.create_sheet(title="Emails")
    ws_emails.views.sheetView[0].showGridLines = True

    email_headers = [
        "Email", "Name", "Role", "Department", "Type", "Purpose",
        "Domain", "Source URL", "Source Title", "Source Type",
        "Occurrences", "Validation Status", "Domain Match", "Confidence"
    ]
    ws_emails.append(email_headers)

    for item in results:
        ws_emails.append([
            item.email,
            item.name or "",
            item.role or "",
            item.department or "",
            item.email_type,
            item.purpose or "",
            item.domain,
            item.source_url,
            item.source_title,
            item.source_type,
            item.occurrences,
            item.validation_status,
            item.domain_match,
            item.confidence
        ])

    _format_sheet(ws_emails, header_font, header_fill, thin_border, hyperlink_col_idx=8)

    # 2. Scan Summary Sheet
    ws_summary = wb.create_sheet(title="Scan Summary")
    ws_summary.views.sheetView[0].showGridLines = True

    ws_summary.append(["Metric", "Value"])
    summary_rows = [
        ("Target Domain", stats.domain),
        ("Scan Start Time", stats.start_time),
        ("Scan Finish Time", stats.finish_time),
        ("Duration (seconds)", stats.scan_duration_seconds),
        ("Pages Discovered", stats.pages_discovered),
        ("Pages Scanned", stats.pages_scanned),
        ("HTML Pages Scanned", stats.html_pages),
        ("PDF Files Processed", stats.pdf_files),
        ("Raw Email Occurrences", stats.raw_email_occurrences),
        ("Unique Emails", stats.unique_emails),
        ("Target Domain Emails", stats.target_domain_emails),
        ("Sites Discovered", stats.sites_discovered),
        ("External Sources Scanned", stats.external_sources_scanned),
        ("Well-Known Record Emails", stats.wellknown_emails_found),
        ("Total Errors Logged", stats.error_count),
    ]

    for metric, val in summary_rows:
        ws_summary.append([metric, val])

    _format_sheet(ws_summary, header_font, header_fill, thin_border)

    # 3. Errors Sheet
    ws_errors = wb.create_sheet(title="Errors")
    ws_errors.views.sheetView[0].showGridLines = True

    ws_errors.append(["URL", "HTTP Status", "Error Type", "Error Message"])
    for err in errors:
        ws_errors.append([
            err.url,
            err.status_code if err.status_code is not None else "N/A",
            err.error_type,
            err.message
        ])

    _format_sheet(ws_errors, header_font, header_fill, thin_border, hyperlink_col_idx=1)

    # Save to buffer
    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)

    return output_buffer.getvalue()


def _format_sheet(
    ws,
    header_font: Font,
    header_fill: PatternFill,
    border: Border,
    hyperlink_col_idx: int = -1
) -> None:
    """Applies headers, auto filters, column widths, freeze panes, and hyperlinks."""
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Style header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = HEADER_ALIGNMENT

    link_font = LINK_FONT

    # Format data rows & set auto column widths
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border

            # Convert URL column to clickable hyperlink
            if col_idx == hyperlink_col_idx and cell.value and str(cell.value).startswith("http"):
                cell.hyperlink = str(cell.value)
                cell.font = link_font

    # Adjust column width automatically
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)
