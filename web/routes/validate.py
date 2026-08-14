import asyncio
import csv
import io
import json
from typing import List, Set

import openpyxl
from fastapi import APIRouter, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from src.utils.logging import logger
from src.utils.validation_store import (
    load_validation_results,
    new_validation_id,
    read_validation_log,
    save_validation_run,
)
from src.validation.domain_validator import get_email_verification
from src.validation.email_validator import normalize_email_address
from web import settings
from web.jobs.batch import BatchJob, validation_registry

router = APIRouter(prefix="/api", tags=["validate"])

# asyncio holds only a weak reference to a running task, so a job with nothing
# else pointing at it can be collected mid-flight. Anchoring it here until it
# finishes is the documented way to keep that from happening.
_background_tasks: Set[asyncio.Task] = set()

_UPLOAD_CHUNK_BYTES = 1024 * 1024


class ValidationResult(BaseModel):
    original_email: str
    normalized_email: str
    is_valid_syntax: bool
    mx_status: str
    mailbox_status: str
    reason: str


def _extract_emails_from_text_block(text: str) -> List[str]:
    emails = []
    for part in str(text).replace(',', ' ').replace(';', ' ').split():
        norm = normalize_email_address(part)
        if norm:
            emails.append(norm)
    return emails


async def _read_capped(file: UploadFile) -> bytes:
    """
    Reads an upload, refusing it as soon as it crosses the size ceiling.

    Chunked rather than a single read() so an oversized file is rejected on the
    way in — reading first and checking afterwards would mean the request that
    is too large to accept is already resident by the time we say so.
    """
    limit = settings.MAX_UPLOAD_BYTES
    chunks: List[bytes] = []
    total = 0

    while True:
        chunk = await file.read(_UPLOAD_CHUNK_BYTES)
        if not chunk:
            break
        total += len(chunk)
        if total > limit:
            raise HTTPException(
                status_code=413,
                detail=f"File exceeds the {limit // (1024 * 1024)}MB upload limit.",
            )
        chunks.append(chunk)

    return b"".join(chunks)


def _notify(job: BatchJob, payload) -> None:
    """Pushes one NDJSON line, or None as the end-of-stream token, to every listener."""
    for queue in list(job.listeners):
        try:
            queue.put_nowait(payload)
        except Exception:
            pass


async def run_validation_job(job: BatchJob, email_list: List[str]) -> None:
    """Background worker task for email batch validation."""
    batch_size = 10

    try:
        for i in range(0, len(email_list), batch_size):
            batch = email_list[i:i + batch_size]

            def process_batch(b):
                res = []
                for em in b:
                    verification = get_email_verification(em, probe_smtp=True)
                    res.append(ValidationResult(
                        original_email=em,
                        normalized_email=em,
                        is_valid_syntax=verification.syntax_valid,
                        mx_status=verification.mx_status,
                        mailbox_status=verification.mailbox_status,
                        reason=verification.provider_notes if verification.provider_notes else ""
                    ))
                return res

            batch_results = await asyncio.to_thread(process_batch, batch)
            dict_batch = [r.model_dump() for r in batch_results]
            processed = job.extend(dict_batch)

            _notify(job, json.dumps({
                "results": dict_batch,
                "processed": processed,
                "total": len(email_list),
            }) + "\n")

        job.mark_terminal("completed")
        await asyncio.to_thread(save_validation_run, job.job_id, job.label, job.results)
    except Exception as exc:
        job.error = str(exc)
        job.mark_terminal("failed")
        logger.error(f"Background validation '{job.job_id}' error: {exc}")
        results = job.results
        if results:
            await asyncio.to_thread(save_validation_run, job.job_id, job.label, results)
    finally:
        job.finished = True
        _notify(job, None)


@router.post("/validate_file")
async def validate_file(file: UploadFile = File(...)):
    """Upload a CSV or Excel file containing emails to validate them in background."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file uploaded")

    if not validation_registry.has_capacity():
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail=(
                f"Validation capacity is full "
                f"({validation_registry.max_concurrent} running). Try again shortly."
            ),
        )

    filename = file.filename
    fname_lower = filename.lower()
    if not fname_lower.endswith((".csv", ".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are supported")

    content = await _read_capped(file)

    emails = set()
    try:
        if fname_lower.endswith(".csv"):
            text = content.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(text))
            for row in reader:
                for cell in row:
                    if cell.strip():
                        emails.update(_extract_emails_from_text_block(cell))
        else:
            wb = openpyxl.load_workbook(filename=io.BytesIO(content), data_only=True, read_only=True)
            try:
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    for row in sheet.iter_rows(values_only=True):
                        for cell in row:
                            if cell and str(cell).strip():
                                emails.update(_extract_emails_from_text_block(str(cell)))
            finally:
                wb.close()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}") from e

    val_id = new_validation_id(filename)
    email_list = list(emails)

    job = BatchJob(job_id=val_id, kind="validation", label=filename, total=len(email_list))
    validation_registry.add(job)

    # Launch background execution independent of the HTTP stream, so a client that
    # disconnects mid-upload does not abandon the work it already paid for.
    task = asyncio.create_task(run_validation_job(job, email_list))
    _background_tasks.add(task)
    task.add_done_callback(_background_tasks.discard)

    async def generate_results():
        q: asyncio.Queue = asyncio.Queue()
        job.listeners.add(q)
        try:
            yield json.dumps({
                "total": len(email_list),
                "validation_id": val_id,
                "processed": job.processed,
                "results": job.results,
            }) + "\n"
            while not job.finished:
                msg = await q.get()
                if msg is None:
                    break
                yield msg
        finally:
            job.listeners.discard(q)

    return StreamingResponse(generate_results(), media_type="application/x-ndjson")


class ExportRequest(BaseModel):
    results: List[ValidationResult]
    format: str


def _build_export_stream(
    results: List[ValidationResult],
    format: str,
    filename_prefix: str = "validation_results",
):
    valid_results = [
        r for r in results
        if r.is_valid_syntax and r.mx_status == "valid" and r.mailbox_status == "valid"
    ]

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Email", "Syntax", "MX Status", "Mailbox Status", "Remarks"])
        for r in results:
            writer.writerow([
                r.original_email, r.is_valid_syntax, r.mx_status, r.mailbox_status, r.reason,
            ])

        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename_prefix}.csv"}
        )
    else:
        wb = openpyxl.Workbook()

        # Sheet 1: All Emails
        ws_all = wb.active
        ws_all.title = "All Emails"
        headers = ["Email", "Syntax", "MX Status", "Mailbox Status", "Remarks"]
        ws_all.append(headers)
        for r in results:
            ws_all.append([
                r.original_email, r.is_valid_syntax, r.mx_status, r.mailbox_status, r.reason,
            ])

        # Sheet 2: Validated Emails
        ws_valid = wb.create_sheet(title="Validated Emails")
        ws_valid.append(headers)
        for r in valid_results:
            ws_valid.append([
                r.original_email, r.is_valid_syntax, r.mx_status, r.mailbox_status, r.reason,
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename_prefix}.xlsx"}
        )


@router.post("/validate_export")
async def validate_export(req: ExportRequest):
    if req.format not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="Invalid format")
    # The rows come from the client, and building a workbook out of them costs
    # memory proportional to however many were sent.
    if len(req.results) > settings.MAX_EXPORT_ROWS:
        raise HTTPException(
            status_code=413,
            detail=f"Export is limited to {settings.MAX_EXPORT_ROWS} rows.",
        )
    return _build_export_stream(req.results, req.format)


@router.get("/validate_history")
async def get_validate_history():
    history_logs = read_validation_log()
    log_map = {item["id"]: item for item in history_logs}

    active_items = []
    for job in validation_registry.all():
        if job.job_id not in log_map:
            active_items.append({
                "id": job.job_id,
                "filename": job.label,
                "total_emails": job.total,
                "created_at": job.created_at.isoformat(),
                "status": job.status,
                "has_results": True,
            })
        else:
            log_map[job.job_id]["status"] = job.status

    for item in history_logs:
        if "status" not in item:
            item["status"] = "completed"

    return active_items + history_logs


@router.get("/validate_history/{validation_id}/results")
async def get_historical_validation_results(validation_id: str):
    job = validation_registry.get(validation_id)
    if job:
        return {
            "id": validation_id,
            "filename": job.label,
            "total_emails": job.total,
            "created_at": job.created_at.isoformat(),
            "status": job.status,
            "results": job.results,
        }
    data = load_validation_results(validation_id)
    if not data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Validation record not found"
        )
    return data


@router.get("/validate_history/{validation_id}/export.{format}")
async def export_historical_validation(validation_id: str, format: str):
    if format not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="Invalid format")

    data = load_validation_results(validation_id)
    if not data:
        job = validation_registry.get(validation_id)
        if job:
            data = {"results": job.results}

    if not data or "results" not in data:
        raise HTTPException(status_code=404, detail="Validation data not found")

    results = [ValidationResult(**item) for item in data["results"]]
    clean_prefix = f"validation_{validation_id}"
    return _build_export_stream(results, format, filename_prefix=clean_prefix)


# Re-exported for the tests, which drive job lifecycle directly.
__all__ = ["router", "ValidationResult", "run_validation_job"]
